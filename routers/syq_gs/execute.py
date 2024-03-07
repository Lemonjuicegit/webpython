import re,asyncio
import pandas as pd
import openpyxl
from pathlib import Path
from openpyxl import load_workbook
from docxtpl import DocxTemplate
from openpyxl.styles import Alignment, Border, Side

def add_border(cell):
    # 给单元格添加边框
    if type(cell) is openpyxl.cell.cell.Cell:
        cell.border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    elif type(cell) is tuple:
        for cel in cell:
            for ce in cel:
               ce.border = Border(left=Side(style='thin'), right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin')) 

def assignment_cell(cell,text):
    if type(cell) is not openpyxl.cell.cell.Cell:
        raise ValueError(f"传入类型：{type(cell)}，需要类型：{openpyxl.cell.cell.Cell}")
    add_border(cell)
    cell.value = text
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
class informationGS:
    def __init__(self, datapath):
      self.df = pd.read_excel(datapath)
      self.qlrlist = set(self.df['权利人'])
      self.cmlist = {re.search(r'(.+?)镇(.+?)(村|社区)',v).group() for v in self.qlrlist}
      self.count = 0
      
    def araeAch(self,cm,save):
        '''
            公示面积成果
        '''
        temp_path = r"E:\exploitation\webpython\routers\syq_gs\template\集体土地所有权调查成果表.xlsx"
        wb = load_workbook(temp_path)
        sheet = wb['Sheet1']
        def adddata(row):
            self.count += 1
            assignment_cell(sheet[f'A{sheet.max_row+1}'],self.count)
            assignment_cell(sheet[f'B{sheet.max_row}'], row['权利人'])
            assignment_cell(sheet[f'C{sheet.max_row}'], row['坐落'])
            assignment_cell(sheet[f'D{sheet.max_row}'], row['权利类型'])
            assignment_cell(sheet[f'E{sheet.max_row}'], row['面积'])
            assignment_cell(sheet[f'F{sheet.max_row}'], row['农用地'])
            assignment_cell(sheet[f'G{sheet.max_row}'], row['耕地'])
            assignment_cell(sheet[f'H{sheet.max_row}'], row['林地'])
            assignment_cell(sheet[f'I{sheet.max_row}'], row['草地'])
            assignment_cell(sheet[f'J{sheet.max_row}'], row['其他'])
            assignment_cell(sheet[f'K{sheet.max_row}'], row['建设用地'])
            assignment_cell(sheet[f'L{sheet.max_row}'], row['未利用地'])
            add_border(sheet[f'M{sheet.max_row}'])
            
        res = self.df[self.df['权利人'].str.contains(cm)]
        res.apply(adddata,axis=1)
        self.count = 0
        wb.save(save)
        
    async def araeAch_all(self,save):
        asyncio.sleep(0)
        for cm in self.cmlist:
            self.araeAch(cm,Path(save) / f"大足区{cm}集体土地所有权调查成果表.xlsx")
            yield f"大足区{cm}集体土地所有权调查成果表"
        
    def noticeGS(self,data,qlr,save):
        '''
        公示通知
        '''
        formatdata = []
        for _,row in data.iterrows():
            formatdata.append({
                'qlrmc':row['权利人'],
                'zl':row['坐落'],
                'mj':row['面积'],
            })
        temp_path = r"E:\exploitation\webpython\routers\syq_gs\template\集体土地公示.docx"
        doc = DocxTemplate(temp_path)
        doc.render({
            'data':formatdata,
            'QLR':qlr
        })
        doc.save(save)
    def noticeGS_all(self,save):
        for cm in self.cmlist:
            self.noticeGS(self.df[self.df['权利人'].str.contains(cm)],cm,Path(save) / f"大足区{cm}集体土地公示.docx")
            yield f"大足区{cm}集体土地公示"

if __name__ == "__main__":
    gs = informationGS(r"E:\工作文档\大足所有权\大足-古龙镇\模板及台账\大足区集体土地所有权分类面积表台账-合并2166.xlsx")
    save_path = r"E:\工作文档\大足所有权\大足-古龙镇\公示\公示公告\新建文件夹"
    gs.araeAch_all(save_path)
    