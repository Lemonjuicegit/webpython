import openpyxl
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
def add_border(cell):
    # 给单元格添加边框
    if type(cell) is openpyxl.cell.cell.Cell:
        cell.border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    elif type(cell) is tuple:
        for cel in cell:
            for ce in cel:
               ce.border = Border(left=Side(style='thin'), right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin')) 

def head_zddm(ws,text):
    ws[f'A{ws.max_row+1}'] = text
    ws[f'A{ws.max_row}'].alignment = Alignment(vertical='center')
    ws.row_dimensions[ws.max_row].height = 26
    add_border(ws[f'A{ws.max_row}'])
    ws.merge_cells(start_row=ws.max_row,start_column=1,end_row=ws.max_row,end_column=ws.max_column)
    assignment_cell(ws[f'A{ws.max_row+1}'],'界址点号')
    assignment_cell(ws[f'B{ws.max_row}'],'距离')
    assignment_cell(ws[f'C{ws.max_row}'],'X坐标')
    assignment_cell( ws[f'D{ws.max_row}'],'Y坐标')
    assignment_cell(ws[f'E{ws.max_row}'],'备注')
    add_border(ws[f'A{ws.max_row}:E{ws.max_row}'])

def assignment_cell(cell,text):
    if type(cell) is not openpyxl.cell.cell.Cell:
        raise ValueError(f"传入类型：{type(cell)}，需要类型：{openpyxl.cell.cell.Cell}")
    cell.value = text
    cell.alignment = Alignment(horizontal='center', vertical='center')

def generate_jzdcg(data,jzd_data, qlr,mj,save_path):
    if not data:
        return
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 17.5
    ws.column_dimensions['B'].width = 17.5
    ws.column_dimensions['C'].width = 17.5
    ws.column_dimensions['D'].width = 17.5
    ws.column_dimensions['E'].width = 17.5
    assignment_cell(ws['A1'],'界址点坐标成果表')
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=5)
    ws['A4'] = f"权利人：{qlr}"
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=3)
    ws['D4'] = f"面积（㎡）:{mj}"
    ws.row_dimensions[4].height = 26
    ws['A4'].alignment = Alignment(vertical='center')
    ws['D4'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=3)
    ws.merge_cells(start_row=4, start_column=4, end_row=4, end_column=5)
    for zddm in data:
        n = 0
        zd_jzd = jzd_data[jzd_data['ZDDM'] == zddm]
        # a= ws.max_row//62
        # b = (zd_jzd.shape[0] + 4 + ws.max_row)//62
        # if (a != b) and ((zd_jzd.shape[0]+ 4 + ws.max_row) != (a+1)*62):
        #     for _ in range((a+1)*62-ws.max_row):
        #         ws[f'A{ws.max_row+1}'].value = ''
        head_zddm(ws,f"宗地代码：{zddm}") 
        for _,row in zd_jzd.iterrows():
            assignment_cell(ws[f'A{ws.max_row+1}'],row['JZD_NEW'])
            ws.row_dimensions[ws.max_row].height = 5
            assignment_cell(ws[f'C{ws.max_row}'],row.geometry.x)
            assignment_cell(ws[f'D{ws.max_row}'],row.geometry.y)
            row_index = ws.max_row
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index+1, end_column=1)
            ws.merge_cells(start_row=row_index, start_column=3, end_row=row_index+1, end_column=3)
            ws.merge_cells(start_row=row_index, start_column=4, end_row=row_index+1, end_column=4)
            ws.merge_cells(start_row=row_index, start_column=5, end_row=row_index+1, end_column=5)
            if n != 0:
                row_index = ws.max_row
                assignment_cell(ws[f'B{row_index-2}'],'/')
                ws.merge_cells(start_row=row_index-2, start_column=2, end_row=row_index-1, end_column=2)
            row_index = ws.max_row
            add_border(ws[f'A{row_index-2}:E{row_index}'])
            n += 1
        assignment_cell(ws[f'A{ws.max_row+1}'],'J1')
        ws.row_dimensions[ws.max_row].height = 5
        assignment_cell(ws[f'C{ws.max_row}'],zd_jzd[zd_jzd.JZD_NEW == 'J1'].geometry.x.values[0])
        assignment_cell(ws[f'D{ws.max_row}'],zd_jzd[zd_jzd.JZD_NEW == 'J1'].geometry.y.values[0])
        row_index = ws.max_row
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index+1, end_column=1)
        ws.merge_cells(start_row=row_index, start_column=3, end_row=row_index+1, end_column=3)
        ws.merge_cells(start_row=row_index, start_column=4, end_row=row_index+1, end_column=4)
        ws.merge_cells(start_row=row_index, start_column=5, end_row=row_index+1, end_column=5)
        row_index = ws.max_row
        assignment_cell(ws[f'B{row_index-2}'],'/')
        ws.merge_cells(start_row=row_index-2, start_column=2, end_row=row_index-1, end_column=2)
        row_index = ws.max_row
        add_border(ws[f'A{row_index-2}:E{row_index}'])
    ws.cell(2, ws.max_row-1).number_format = '0.000'
    ws.cell(3, ws.max_row-1).number_format = '0.000'
    assignment_cell(ws[f'A{ws.max_row+1}'],'计算者：')
    assignment_cell(ws[f'B{ws.max_row}'],'袁珊')
    assignment_cell(ws[f'D{ws.max_row}'],'查核者')
    assignment_cell(ws[f'E{ws.max_row}'],'余勇')
    
    wb.save(save_path)
    

def get_jzd_data(jzd_data,zd_data):
    jzd_data.sort_values(by='ZDDM', inplace=True)
    jzd_data.sort_values(by='PX', inplace=True)
    data = {}
    for _,row in zd_data.iterrows():
        if row['QLRMC'] not in data:
            if len(jzd_data[jzd_data['ZDDM']==row['ZDDM']]) == 0:
                data[row['QLRMC']] = {'SETZDDM':set(),'ZDMJ':row['ZDMJ']}
                raise Exception(f"{row['ZDDM']}:没有界址点")
            data[row['QLRMC']] = {'SETZDDM':{row['ZDDM']},'ZDMJ':row['ZDMJ']}
        else:
            if len(jzd_data[jzd_data['ZDDM']==row['ZDDM']]) == 0:
                data[row['QLRMC']] = {'SETZDDM':set(),'ZDMJ':row['ZDMJ']}
                raise Exception(f"{row['ZDDM']}:没有界址点")
            data[row['QLRMC']]['SETZDDM'].add(row['ZDDM'])
            data[row['QLRMC']]['ZDMJ'] = data[row['QLRMC']]['ZDMJ'] + row['ZDMJ']
    return data

def generate_jzdcg_all(jzd_data,zd_data,save_path):
    data = get_jzd_data(jzd_data,zd_data)
    for qlr,value in data.items():
        if not value['SETZDDM']:
            yield f"{qlr}没有界址点成果表"
        generate_jzdcg(value['SETZDDM'],jzd_data,qlr,round(value['ZDMJ'],4),Path(save_path)/f"{qlr}界址点成果表.xlsx")
        yield f"{qlr}界址点成果表"

    


'''
BaseException	所有异常的基类
SystemExit	解释器请求退出
KeyboardInterrupt	用户中断执行(通常是输入^C)
Exception	常规错误的基类
StopIteration	迭代器没有更多的值
GeneratorExit	生成器(generator)发生异常来通知退出
StandardError	所有的内建标准异常的基类
ArithmeticError	所有数值计算错误的基类
FloatingPointError	浮点计算错误
OverflowError	数值运算超出最大限制
ZeroDivisionError	除(或取模)零 (所有数据类型)
AssertionError	断言语句失败
AttributeError	对象没有这个属性
EOFError	没有内建输入,到达EOF 标记
EnvironmentError	操作系统错误的基类
IOError	输入/输出操作失败
OSError	操作系统错误
WindowsError	系统调用失败
ImportError	导入模块/对象失败
LookupError	无效数据查询的基类
IndexError	序列中没有此索引(index)
KeyError	映射中没有这个键
MemoryError	内存溢出错误(对于Python 解释器不是致命的)
NameError	未声明/初始化对象 (没有属性)
UnboundLocalError	访问未初始化的本地变量
ReferenceError	弱引用(Weak reference)试图访问已经垃圾回收了的对象
RuntimeError	一般的运行时错误
NotImplementedError	尚未实现的方法
SyntaxError	Python 语法错误
IndentationError	缩进错误
TabError	Tab 和空格混用
SystemError	一般的解释器系统错误
TypeError	对类型无效的操作
ValueError	传入无效的参数
UnicodeError	Unicode 相关的错误
UnicodeDecodeError	Unicode 解码时的错误
UnicodeEncodeError	Unicode 编码时错误
UnicodeTranslateError	Unicode 转换时错误
Warning	警告的基类
DeprecationWarning	关于被弃用的特征的警告
FutureWarning	关于构造将来语义会有改变的警告
OverflowWarning	旧的关于自动提升为长整型(long)的警告
PendingDeprecationWarning	关于特性将会被废弃的警告
RuntimeWarning	可疑的运行时行为(runtime behavior)的警告
SyntaxWarning	可疑的语法的警告
UserWarning	用户代码生成的警告
'''
